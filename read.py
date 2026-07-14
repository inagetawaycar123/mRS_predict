from openpyxl import load_workbook
import numpy as np
import sys

try:
    from sklearn.preprocessing import Standar_90_MRScaler, LabelEncoder
    from sklearn.svm import SVC
    from sklearn.model_selection import train_test_split
    from sklearn.metrics import classification_report, accuracy_score
    import joblib
except Exception as e:
    print('请先安装 scikit-learn 和 joblib：pip install scikit-learn joblib')
    raise

# 加载工作簿并选择工作表（优先使用名为 '1_原始Master' 的表）
wb = load_workbook('ProveIt_中文翻译_编码保留版.xlsx', read_only=True)
sheet_name = '1_原始Master' if '1_原始Master' in wb.sheetnames else wb.sheetnames[0]
ws = wb[sheet_name]

# 读取表头，定位所需列（支持若干常见命名变体）
headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
print('Headers:', headers)

def find_col(headers, candidates):
    low_headers = [ (h.lower() if h is not None else '') for h in headers ]
    for cand in candidates:
        cand_low = cand.lower()
        # 首先尝试精确匹配
        if cand_low in low_headers:
            return low_headers.index(cand_low)
    # 再尝试包含匹配
    for cand in candidates:
        cand_low = cand.lower()
        for i,h in enumerate(low_headers):
            if cand_low in h and h!='':
                return i
    return None

# 定义候选列名（根据你的数据可能需调整）
CTP_core_candidates = ['CTP_core_rCBF_thresh12.5', 'CTP_core', 'CTP_core_rCBF', 'CTP_core_rCBF_thresh']
_90_MRS_candidates = ['90 Day MRS', '90day mrs', 'mRS', '90 day mrs']
gender_candidates = ['Gender', 'Sex', 'gender', 'sex']
age_candidates = ['Age', 'age']
onset_candidates = ['Onset to CT time', 'Onset to CT', 'Onset_to_CT', 'Onset time', 'Onset']
nihss_candidates = ['NIHSS Baseline', 'NIHSS_baseline', 'NIHSS', 'NIHSS baseline']

CTP_core_idx = find_col(headers, CTP_core_candidates)
_90_MRS_idx = find_col(headers, _90_MRS_candidates)
gender_idx = find_col(headers, gender_candidates)
age_idx = find_col(headers, age_candidates)
onset_idx = find_col(headers, onset_candidates)
nihss_idx = find_col(headers, nihss_candidates)

if CTP_core_idx is None or _90_MRS_idx is None:
    print('未找到 CTP_core 或 _90_MRS 列，请检查表头是否包含这两个字段（候选名：', CTP_core_candidates, _90_MRS_candidates, ')')
    sys.exit(1)

print('列索引：CTP_core=%s, _90_MRS=%s, Gender=%s, Age=%s, Onset=%s, NIHSS=%s' % (CTP_core_idx, _90_MRS_idx, gender_idx, age_idx, onset_idx, nihss_idx))

def to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        try:
            s = str(v).strip()
            # 移除非数字字符
            import re
            s2 = re.sub('[^0-9.+-eE]', '', s)
            return float(s2) if s2!='' else None
        except Exception:
            return None

def parse_gender(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip().lower()
    if s in ('m','male','男','1','man'):
        return 1
    if s in ('f','female','女','0','woman'):
        return 0
    return None

# 收集数据：特征顺序为 [CTP_core, Gender, Age, OnsetToCT, NIHSS]
X_list = []
y_list = []
for row in ws.iter_rows(min_row=2, values_only=True):
    CTP_core = row[CTP_core_idx]
    _90_MRS = row[_90_MRS_idx]
    # 必须同时有 CTP_core 和 _90_MRS
    if CTP_core is None or _90_MRS is None:
        continue
    CTP_core_val = to_float(CTP_core)
    if CTP_core_val is None:
        continue

    # 其它特征（可缺失，但我们这里要求全部存在）
    gender_val = parse_gender(row[gender_idx]) if gender_idx is not None else None
    age_val = to_float(row[age_idx]) if age_idx is not None else None
    onset_val = to_float(row[onset_idx]) if onset_idx is not None else None
    nihss_val = to_float(row[nihss_idx]) if nihss_idx is not None else None

    # 若你希望允许部分缺失可放宽此条件；当前要求所有四个额外字段不为空
    if gender_val is None or age_val is None or onset_val is None or nihss_val is None:
        continue

    X_list.append([CTP_core_val, gender_val, age_val, onset_val, nihss_val])
    y_list.append(_90_MRS)

if len(X_list) == 0:
    print('没有可用的数据行（满足所有特征且 _90_MRS 存在）。')
    sys.exit(1)

X = np.array(X_list)
y_raw = np.array(y_list)

# 将标签转换为整数并过滤 0-6
y_converted = []
for val in y_raw:
    try:
        num = int(float(val))
        y_converted.append(num)
    except Exception:
        y_converted.append(None)
y = np.array(y_converted, dtype=object)

allowed = set(range(7))
mask_allowed = np.array([ (v is not None and v in allowed) for v in y ])
X = X[mask_allowed]
y = y[mask_allowed].astype(int)

if len(X) == 0:
    print('过滤后没有可用样本（仅保留 _90_MRS=0..6）。')
    sys.exit(1)

# 移除样本数少于 2 的类别
unique, counts = np.unique(y, return_counts=True)
labels_to_keep = [lab for lab, cnt in zip(unique, counts) if cnt >= 2]
removed_labels = [int(lab) for lab, cnt in zip(unique, counts) if cnt < 2]
if removed_labels:
    print('以下标签样本过少，将被移除：', removed_labels)
    mask_counts = np.isin(y, labels_to_keep)
    X = X[mask_counts]
    y = y[mask_counts]

if len(X) == 0:
    print('移除样本数少的类别后无可用样本。')
    sys.exit(1)

# 如果标签仍然为非数值类型（一般不会出现），使用 LabelEncoder
le = None
if y.dtype.kind not in 'iuf':
    le = LabelEncoder()
    y = le.fit_transform(y)

# 根据现有类别情况决定是否使用 stratify
unique_labels = np.unique(y)
stratify_arg = y if (len(unique_labels) > 1 and np.min(np.bincount(y)) >= 2) else None
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42, stratify=stratify_arg)

# 标准化特征
scaler = Standar_90_MRScaler()
X_train_s = scaler.fit_transform(X_train)
X_test_s = scaler.transform(X_test)

# 训练 SVM（用于分类）
clf = SVC(kernel='rbf', C=1.0, probability=False, random_state=42)
clf.fit(X_train_s, y_train)

# 评估
y_pred = clf.predict(X_test_s)
print('Accuracy:', accuracy_score(y_test, y_pred))
print('Classification report:\n', classification_report(y_test, y_pred))

# 保存模型、scaler 和 label encoder（如果有）
model_bundle = {'model': clf, 'scaler': scaler}
if le is not None:
    model_bundle['label_encoder'] = le
joblib.dump(model_bundle, 'svm_model_bundle.joblib')
print('模型已保存为 svm_model_bundle.joblib')