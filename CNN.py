from pathlib import Path
from collections import defaultdict
import re
import sys

import numpy as np
from openpyxl import load_workbook

try:
    import torch
    import torch.nn as nn
    from torch.utils.data import Dataset, DataLoader
    from sklearn.model_selection import train_test_split
    from sklearn.metrics import classification_report, accuracy_score
    import joblib
except Exception:
    print('请先安装依赖：pip install torch scikit-learn joblib openpyxl')
    raise


def patient_sort_key(pid: str):
    return tuple(int(x) if x.isdigit() else x for x in re.split(r'(\d+)', pid) if x != '')


def normalize_patient_id(pid):
    if pid is None:
        return None
    text = str(pid).strip()
    if text == '':
        return None
    if text.lower().startswith('prove-it-'):
        return text[len('Prove-IT-'):]
    return text


def parse_slice_name(file_path: Path):
    stem = file_path.stem
    if '_' not in stem:
        raise ValueError(f'文件名不符合 病人ID_切片编号.npy 格式: {file_path.name}')
    pid, slice_no = stem.rsplit('_', 1)
    try:
        slice_idx = int(slice_no)
    except ValueError as exc:
        raise ValueError(f'切片编号不是整数: {file_path.name}') from exc
    return pid, slice_idx


def find_col(headers, candidates):
    low_headers = [(h.lower() if h is not None else '') for h in headers]
    for cand in candidates:
        cand_low = cand.lower()
        if cand_low in low_headers:
            return low_headers.index(cand_low)
    for cand in candidates:
        cand_low = cand.lower()
        for i, h in enumerate(low_headers):
            if cand_low in h and h != '':
                return i
    return None

def load_patient_slices(npy_dir='np_a_normalized'): 
    npy_path = Path(npy_dir) 
    if not npy_path.exists() or not npy_path.is_dir(): 
        print(f'未找到数据目录: {npy_dir}') 
        sys.exit(1) 
        
    patient_slices = defaultdict(list) 
    for file_path in npy_path.glob('*.npy'): 
        pid, slice_idx = parse_slice_name(file_path) 
        patient_slices[pid].append((slice_idx, file_path)) 
        
    if len(patient_slices) == 0: 
        print(f'{npy_dir} 中没有找到任何 .npy 文件') 
        sys.exit(1) 
    patient_ids = sorted(patient_slices.keys(), key=patient_sort_key) 
    ordered = {} 
    for pid in patient_ids: 
        ordered[pid] = [p for _, p in sorted(patient_slices[pid], key=lambda x: x[0])] 
    return patient_ids, ordered

def load_patient_labels(excel_path):

    wb = load_workbook(excel_path, read_only=True)

    sheet_name = '1_原始Master' if '1_原始Master' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]


    rows = ws.iter_rows(values_only=True)

    headers = list(next(rows))


    # 找病人ID列
    pid_idx = find_col(
        headers,
        [
            'Patient ID',
            'PatientID',
            'patient id',
            'ID',
            '编号'
        ]
    )


    # 找mRS列
    mrs_idx = find_col(
        headers,
        [
            '90 Day MRS',
            '90day mrs',
            '90 day mrs',
            'mRS'
        ]
    )


    if pid_idx is None:
        print("Excel中没有找到Patient ID列")
        sys.exit(1)

    if mrs_idx is None:
        print("Excel中没有找到90 Day MRS列")
        sys.exit(1)



    patient_mrs = {}


    for row in rows:

        pid = row[pid_idx]
        mrs = row[mrs_idx]


        if pid is None:
            continue


        pid = normalize_patient_id(pid)
        if pid is None:
            continue


        try:
            mrs = int(float(mrs))
        except:
            continue


        patient_mrs[pid] = mrs


    print("Excel读取患者数量:", len(patient_mrs))


    return patient_mrs

class SliceDataset(Dataset):
    def __init__(self, samples):
        self.samples = samples

    def __len__(self):
        return len(self.samples)

    def __getitem__(self, idx):
        file_path, label = self.samples[idx]
        arr = np.load(file_path)
        arr = np.asarray(arr, dtype=np.float32)

        if arr.ndim == 3:
            # 若是多通道体素切片，压到单通道
            arr = arr.mean(axis=0)
        elif arr.ndim != 2:
            raise ValueError(f'不支持的切片维度: {arr.shape}, file={file_path}')

        # 每张切片做标准化，提升训练稳定性
        mean = float(arr.mean())
        std = float(arr.std())
        if std < 1e-6:
            std = 1.0
        arr = (arr - mean) / std

        x = torch.from_numpy(arr).unsqueeze(0)  # [1, H, W]
        y = torch.tensor(label, dtype=torch.long)
        return x, y


class SimpleCNN(nn.Module):
    def __init__(self, num_classes=7):
        super().__init__()
        self.features = nn.Sequential(
            nn.Conv2d(1, 16, kernel_size=3, padding=1),
            nn.ReLU(inplace=True),
            nn.MaxPool2d(2),
            nn.Conv2d(16, 32, kernel_size=3, padding=1),
            nn.ReLU(inplace=True),
            nn.MaxPool2d(2),
            nn.Conv2d(32, 64, kernel_size=3, padding=1),
            nn.ReLU(inplace=True),
            nn.AdaptiveAvgPool2d((1, 1)),
        )
        self.classifier = nn.Linear(64, num_classes)

    def forward(self, x):
        x = self.features(x)
        x = x.flatten(1)
        return self.classifier(x)


def build_slice_samples(
        patient_ids,
        patient_slices,
        patient_mrs,
        allowed_labels
):


    valid_patient_ids = []
    valid_y = []


    for pid in patient_ids:


        # Excel没有这个病人
        if pid not in patient_mrs:
            continue


        y = patient_mrs[pid]


        if y not in allowed_labels:
            continue


        if pid not in patient_slices:
            continue


        if len(patient_slices[pid]) == 0:
            continue


        valid_patient_ids.append(pid)
        valid_y.append(y)



    if len(valid_patient_ids)==0:
        print("没有有效患者")
        sys.exit(1)



    valid_y=np.array(valid_y)



    # 删除样本太少类别
    unique, counts=np.unique(
        valid_y,
        return_counts=True
    )


    keep_labels=[
        lab for lab,cnt in zip(unique,counts)
        if cnt>=2
    ]



    mask=np.isin(
        valid_y,
        keep_labels
    )


    valid_patient_ids=[
        pid
        for pid,m in zip(valid_patient_ids,mask)
        if m
    ]


    valid_y=valid_y[mask]



    print(
        "有效患者:",
        len(valid_patient_ids)
    )


    print(
        "类别分布:",
        np.unique(
            valid_y,
            return_counts=True
        )
    )



    stratify_arg=None

    if len(np.unique(valid_y))>1:
        stratify_arg=valid_y



    train_pids,test_pids,y_train,y_test=train_test_split(
        valid_patient_ids,
        valid_y,
        test_size=0.2,
        random_state=42,
        stratify=stratify_arg
    )



    train_samples=[]
    test_samples=[]



    for pid,y in zip(train_pids,y_train):

        for slice_path in patient_slices[pid]:

            train_samples.append(
                (
                    slice_path,
                    int(y)
                )
            )



    for pid,y in zip(test_pids,y_test):

        for slice_path in patient_slices[pid]:

            test_samples.append(
                (
                    slice_path,
                    int(y)
                )
            )



    return train_samples,test_samples

def train_model(model, train_loader, device, epochs=10, lr=1e-3):
    criterion = nn.CrossEntropyLoss()
    optimizer = torch.optim.Adam(model.parameters(), lr=lr)

    model.train()
    for epoch in range(1, epochs + 1):
        running_loss = 0.0
        total = 0
        correct = 0
        for x, y in train_loader:
            x = x.to(device)
            y = y.to(device)

            optimizer.zero_grad()
            logits = model(x)
            loss = criterion(logits, y)
            loss.backward()
            optimizer.step()

            running_loss += loss.item() * x.size(0)
            pred = logits.argmax(dim=1)
            correct += (pred == y).sum().item()
            total += x.size(0)

        avg_loss = running_loss / max(total, 1)
        acc = correct / max(total, 1)
        print(f'Epoch {epoch:02d}/{epochs} - loss: {avg_loss:.4f} - acc: {acc:.4f}')


def evaluate_model(model, test_loader, device):
    model.eval()
    y_true = []
    y_pred = []

    with torch.no_grad():
        for x, y in test_loader:
            x = x.to(device)
            logits = model(x)
            pred = logits.argmax(dim=1).cpu().numpy()
            y_pred.extend(pred.tolist())
            y_true.extend(y.numpy().tolist())

    print('Accuracy:', accuracy_score(y_true, y_pred))
    print('Classification report:\n', classification_report(y_true, y_pred, zero_division=0))


def main():
    np.random.seed(42)
    torch.manual_seed(42)

    patient_ids, patient_slices = load_patient_slices(
    'np_a_normalized'
    )
    patient_mrs = load_patient_labels(
        'ProveIt_中文翻译_编码保留版.xlsx'
    )

    train_samples,test_samples = build_slice_samples(
    patient_ids,
    patient_slices,
    patient_mrs,
    set(range(7))
    )

    print('训练切片数:', len(train_samples), '测试切片数:', len(test_samples))

    train_dataset = SliceDataset(train_samples)
    test_dataset = SliceDataset(test_samples)

    train_loader = DataLoader(train_dataset, batch_size=16, shuffle=True, num_workers=0)
    test_loader = DataLoader(test_dataset, batch_size=16, shuffle=False, num_workers=0)

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print('Device:', device)

    model = SimpleCNN(num_classes=7).to(device)
    train_model(model, train_loader, device, epochs=20, lr=1e-3)
    evaluate_model(model, test_loader, device)

    torch.save(model.state_dict(), 'cnn_model.pth')
    joblib.dump({'num_classes': 7}, 'cnn_meta.joblib')
    print('模型已保存: cnn_model.pth, cnn_meta.joblib')


if __name__ == '__main__':
    main()
